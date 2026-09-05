# ─────────────────────────────────────────────
#  utils.py  –  Shared utilities
#  Used by realtime_processor.py and batch_processor.py
# ─────────────────────────────────────────────

import base64
import hashlib
import io
import json
import re
import time
import traceback
from datetime import datetime

import cv2
import numpy as np
import resend

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageChops

try:
    import pytesseract
except ImportError:
    pytesseract = None

import config


# ── PDF page counting ─────────────────────────────────────────────────────────

def count_pdf_pages(file) -> dict:
    """Returns the number of pages in an uploaded PDF without extracting text."""
    try:
        file.seek(0)
        with pdfplumber.open(file) as pdf:
            page_count = len(pdf.pages)
        file.seek(0)
        return {"success": True, "page_count": page_count, "error": None}
    except Exception as e:
        try:
            file.seek(0)
        except Exception:
            pass
        return {"success": False, "page_count": 0, "error": str(e)}


def count_uploaded_pdf_pages(uploaded_files: list) -> dict:
    """Counts pages for all uploaded PDFs and returns per-file detail."""
    total_pages = 0
    files = []
    errors = []

    for f in uploaded_files or []:
        result = count_pdf_pages(f)
        files.append({"name": f.name, **result})
        if result["success"]:
            total_pages += result["page_count"]
        else:
            errors.append(f"{f.name}: {result['error']}")

    return {
        "success": not errors,
        "total_pages": total_pages,
        "files": files,
        "errors": errors,
    }


# ── Upload-time duplicate detection ───────────────────────────────────────────

GSTIN_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", re.I)

INVOICE_NO_REJECTS = {"DATE", "DATED", "NO", "NUMBER", "INVOICE", "TAX", "MODE", "TERMS"}

INVOICE_NO_PATTERNS = [
    re.compile(
        r"\b(?:invoice|inv\.?|bill)\s*"
        r"(?:no\.?|number|#)\s*[:\-]?\s*([A-Z0-9][A-Z0-9/._\-]{1,})",
        re.I,
    ),
    re.compile(r"\b(?:invoice|inv\.?)\s*#\s*([A-Z0-9][A-Z0-9/._\-]{1,})", re.I),
]


def _normalize_invoice_no(value: str) -> str:
    value = str(value or "").strip().upper()
    value = value.strip(" .,:;#")
    return re.sub(r"\s+", "", value)


def extract_invoice_identity_from_text(text: str) -> dict:
    """
    Extracts a strict duplicate identity from local PDF text.
    Only returns a usable key when both GSTIN and invoice number are found.
    """
    source = text or ""
    gstin_match = GSTIN_RE.search(source)
    gstin = gstin_match.group(0).upper() if gstin_match else ""

    invoice_no = ""
    for pattern in INVOICE_NO_PATTERNS:
        for match in pattern.finditer(source):
            candidate = _normalize_invoice_no(match.group(1))
            if (
                candidate
                and candidate not in INVOICE_NO_REJECTS
                and not GSTIN_RE.fullmatch(candidate)
            ):
                invoice_no = candidate
                break
        if invoice_no:
            break

    key = f"{gstin}|{invoice_no}" if gstin and invoice_no else ""
    return {
        "gstin": gstin,
        "invoice_no": invoice_no,
        "key": key,
        "usable": bool(key),
    }


def detect_duplicate_uploads(uploaded_files: list) -> dict:
    """
    Best-effort upload-time duplicate detection.
    Skips only high-confidence duplicates using vendor GSTIN + invoice number.
    """
    seen = {}
    unique_files = []
    duplicates = []
    unidentified = []
    identities = []

    for f in uploaded_files or []:
        # light=True: this is just an identity precheck (GSTIN + invoice no.
        # from native text) — no need to render/OCR pages with no text layer,
        # and Streamlit's rerun model means this can run several times per
        # user action, so the cheap path matters here more than elsewhere.
        extraction = extract_text_from_pdf(f, light=True)
        identity = extract_invoice_identity_from_text(extraction.get("text", "")) if extraction.get("success") else {
            "gstin": "",
            "invoice_no": "",
            "key": "",
            "usable": False,
        }
        f.seek(0)

        record = {"name": f.name, **identity}
        identities.append(record)

        if not identity["usable"]:
            unique_files.append(f)
            unidentified.append(f.name)
            continue

        if identity["key"] in seen:
            duplicates.append({
                "name": f.name,
                "duplicate_of": seen[identity["key"]],
                "gstin": identity["gstin"],
                "invoice_no": identity["invoice_no"],
            })
            continue

        seen[identity["key"]] = f.name
        unique_files.append(f)

    return {
        "unique_files": unique_files,
        "duplicates": duplicates,
        "unidentified": unidentified,
        "identities": identities,
    }


# ── Local OCR (for pages with no native text layer) ───────────────────────────

_ocr_availability_cache = None


def _ocr_available() -> bool:
    """
    Checks, once, whether local Tesseract OCR can actually run (the Python
    package alone isn't enough — the `tesseract` binary must be installed on
    the host). Cached so this isn't re-checked on every page. If it's not
    available, callers fall back to the pre-OCR behavior of always sending
    scanned pages to Claude as images.
    """
    global _ocr_availability_cache
    if _ocr_availability_cache is None:
        if pytesseract is None:
            _ocr_availability_cache = False
        else:
            try:
                pytesseract.get_tesseract_version()
                _ocr_availability_cache = True
            except Exception:
                _ocr_availability_cache = False
    return _ocr_availability_cache


def _page_has_annotation(pil_image, threshold: float = None) -> bool:
    """
    Detects handwriting/rubber-stamp ink on a page image.

    Pen and stamp ink is almost always colored (blue/purple/red); printed
    text is black or gray. Flags the page as annotated when the fraction of
    colored, non-white pixels exceeds `threshold`
    (config.HANDWRITING_INK_THRESHOLD by default).

    Validated against real scans: clean printed pages measure ~0%, pages
    with handwritten notes or stamps measure 4-8%+. Local OCR has been
    observed to silently misread handwritten numbers (e.g. "9000" as "2000")
    without a low-confidence signal to catch it, so annotated pages skip OCR
    entirely and are sent to Claude as images instead.

    Runs on a downscaled thumbnail (config.ANNOTATION_CHECK_MAX_DIM), not
    the full-resolution render — a colored-pixel fraction doesn't need
    full resolution to be representative, and this is the single biggest
    per-page CPU cost in the fallback pipeline, so shrinking it matters most
    on CPU-constrained hosts. Does not affect the image actually sent to
    Claude — pil_image itself is untouched.
    """
    threshold = config.HANDWRITING_INK_THRESHOLD if threshold is None else threshold
    rgb = pil_image.convert("RGB")
    if rgb.width == 0 or rgb.height == 0:
        return False
    thumb = rgb.copy()
    thumb.thumbnail((config.ANNOTATION_CHECK_MAX_DIM, config.ANNOTATION_CHECK_MAX_DIM))
    hsv = thumb.convert("HSV")
    _, s, v = hsv.split()
    sat_mask         = s.point(lambda p: 255 if p > 30  else 0)
    dark_enough_mask = v.point(lambda p: 255 if p < 240 else 0)
    combined = ImageChops.multiply(sat_mask, dark_enough_mask)
    colored_pixels = combined.histogram()[255]
    return (colored_pixels / (thumb.width * thumb.height)) > threshold


def _ocr_page_text(pil_image) -> str:
    """Runs local Tesseract OCR on a page image. Returns "" on any failure."""
    if not _ocr_available():
        return ""
    try:
        return (pytesseract.image_to_string(pil_image) or "").strip()
    except Exception:
        return ""


def _encode_page_jpeg(pil_image) -> bytes:
    """
    Encodes a rendered page to JPEG bytes immediately after rendering, so the
    much larger decoded PIL Image (~10MB+ per page at OCR_RENDER_RESOLUTION_DPI)
    isn't kept in memory alongside every other fallback page in the same file —
    only the small JPEG bytes are. Matters on memory-constrained hosts: a file
    with many fallback pages previously held all of them decoded at once.
    """
    buf = io.BytesIO()
    pil_image.convert("RGB").save(buf, format="JPEG", quality=config.IMAGE_JPEG_QUALITY)
    return buf.getvalue()


# ── Document auto-crop (camera captures only) ─────────────────────────────────
#
#  Adobe Scan-style automatic edge detection: find the document's 4 corners in
#  a captured photo and perspective-warp just that region, dropping whatever
#  background/desk/hand is around it. There's no live camera feed available
#  through st.camera_input() (it only hands back the final snapped photo), so
#  this runs once, right after capture — the corners get drawn on the original
#  photo as a post-capture confirmation, while the actual page sent for
#  extraction is the cropped, flattened version. A low-confidence detection
#  (poor contrast, cluttered background, no clear rectangle) silently falls
#  back to the untouched photo rather than risking a bad crop.

def _order_corners(pts: np.ndarray) -> np.ndarray:
    """Orders 4 points as top-left, top-right, bottom-right, bottom-left."""
    rect = np.zeros((4, 2), dtype="float32")
    total = pts.sum(axis=1)
    rect[0] = pts[np.argmin(total)]
    rect[2] = pts[np.argmax(total)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect


def detect_document_corners(pil_image):
    """
    Looks for a document-shaped quadrilateral in a captured photo. Returns an
    ordered 4x2 float array of (x, y) corners in the original image's
    coordinate space, or None if nothing confident enough was found.
    """
    img = np.array(pil_image.convert("RGB"))
    h, w = img.shape[:2]

    # Downscale for speed — edge detection doesn't need full resolution, and
    # this runs on a single CPU-constrained host.
    scale = 1000.0 / max(h, w) if max(h, w) > 1000 else 1.0
    small = cv2.resize(img, (int(w * scale), int(h * scale))) if scale != 1.0 else img

    gray = cv2.cvtColor(small, cv2.COLOR_RGB2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.dilate(cv2.Canny(blurred, 50, 150), None, iterations=1)

    contours, _ = cv2.findContours(edged, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None

    small_area = small.shape[0] * small.shape[1]
    for contour in sorted(contours, key=cv2.contourArea, reverse=True)[:5]:
        peri = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.02 * peri, True)
        if len(approx) != 4:
            continue
        area_frac = cv2.contourArea(approx) / small_area
        # Lower bound: a small/irrelevant quadrilateral (a corner of a table,
        # a shadow) isn't the document. Upper bound: on a noisy/busy frame,
        # findContours can pick up the image's own border as a "4-point"
        # contour spanning ~all of it — a real photographed document is held
        # with visible background margin around it, so near-full-frame is a
        # detection artifact, not a document.
        if 0.2 < area_frac < 0.92:
            corners = approx.reshape(4, 2).astype("float32") / scale
            return _order_corners(corners)

    return None


def crop_to_document(pil_image, corners):
    """Perspective-warps the image so the given 4 corners become a flat rectangle."""
    img = np.array(pil_image.convert("RGB"))
    tl, tr, br, bl = corners

    max_width = max(int(np.linalg.norm(br - bl)), int(np.linalg.norm(tr - tl)))
    max_height = max(int(np.linalg.norm(tr - br)), int(np.linalg.norm(tl - bl)))
    if max_width < 10 or max_height < 10:
        return pil_image  # degenerate quad — bail out rather than a sliver crop

    dst = np.array([
        [0, 0],
        [max_width - 1, 0],
        [max_width - 1, max_height - 1],
        [0, max_height - 1],
    ], dtype="float32")

    matrix = cv2.getPerspectiveTransform(corners, dst)
    warped = cv2.warpPerspective(img, matrix, (max_width, max_height))
    return Image.fromarray(warped)


def _draw_corner_overlay_jpeg(pil_image, corners) -> bytes:
    """Draws the detected outline on a copy of the photo, for a post-capture
    visual confirmation of what got picked up — not the cropped image itself."""
    img = np.array(pil_image.convert("RGB")).copy()
    pts = corners.astype(int)
    thickness = max(2, img.shape[1] // 250)
    cv2.polylines(img, [pts], isClosed=True, color=(0, 255, 0), thickness=thickness)
    for (x, y) in pts:
        cv2.circle(img, (int(x), int(y)), max(4, img.shape[1] // 150), (0, 255, 0), -1)
    buf = io.BytesIO()
    Image.fromarray(img).save(buf, format="JPEG", quality=config.IMAGE_JPEG_QUALITY)
    return buf.getvalue()


def process_captured_page(img_bytes: bytes) -> dict:
    """
    Runs auto-crop detection on a freshly captured photo. Returns:
        {
            "submit":  bytes  — page to send for extraction (cropped if detected)
            "preview": bytes  — photo for the UI thumbnail (corners drawn on if detected)
            "cropped": bool   — whether a confident detection was found
        }
    Falls back to the untouched original for both "submit" and "preview" when
    detection fails or the captured bytes aren't a decodable image — forcing a
    bad crop is worse than sending the full photo.
    """
    try:
        pil_image = Image.open(io.BytesIO(img_bytes))
        pil_image.load()
    except Exception:
        return {"submit": img_bytes, "preview": img_bytes, "cropped": False}

    try:
        corners = detect_document_corners(pil_image)
    except Exception:
        corners = None

    if corners is None:
        return {"submit": img_bytes, "preview": img_bytes, "cropped": False}

    try:
        preview_bytes = _draw_corner_overlay_jpeg(pil_image, corners)
        cropped_image = crop_to_document(pil_image, corners)
        buf = io.BytesIO()
        cropped_image.convert("RGB").save(buf, format="JPEG", quality=config.IMAGE_JPEG_QUALITY)
        return {"submit": buf.getvalue(), "preview": preview_bytes, "cropped": True}
    except Exception:
        return {"submit": img_bytes, "preview": img_bytes, "cropped": False}


# ── PDF text extraction ───────────────────────────────────────────────────────

def extract_text_from_pdf(file, light: bool = False) -> dict:
    """
    Attempts to extract text from a PDF file using pdfplumber, per page:

      1. Native text layer (fastest, free, most reliable) — used if present.
      2. No text layer: render the page and check for handwriting/stamps.
         - Clean (no annotation) -> run local Tesseract OCR (free); use it if
           it produced enough text.
         - Annotated, or OCR still too sparse, or OCR unavailable -> keep the
           page image to send to Claude directly instead of guessing locally.

    Exact duplicate pages (by content hash) are skipped either way.

    light: when True, skips step 2 entirely for pages with no native text —
    no rendering, no annotation check, no OCR, no fallback image. Pages
    without a text layer are just counted as scanned. Use this for anything
    that only needs native text (e.g. the upload-time duplicate-identity
    precheck) — rendering every page to an image is real CPU cost that isn't
    needed there, and Streamlit's rerun-the-whole-script-on-every-interaction
    model means a caller like that can run many times per user action.

    Returns:
        {
            "success":         bool  — True if there's anything to send (text and/or images)
            "text":            str   — combined text from native-extracted + OCR'd pages
            "page_count":      int   — total pages in PDF
            "skipped_pages":   int   — pages skipped as exact duplicates
            "scanned_pages":   int   — pages sent to Claude as images (annotated, or OCR failed)
            "ocr_pages":       int   — pages recovered via local OCR instead of an image
            "fallback_images": list  — [(page_number, jpeg_bytes), ...] needing image content blocks
            "use_fallback":    bool  — True if any page needs an image (kept for compatibility)
        }
    """
    file.seek(0)
    try:
        seen_hashes      = set()
        pages_text       = []
        scanned_pages    = 0
        ocr_pages        = 0
        skipped_pages    = 0
        total_pages      = 0
        fallback_images  = []

        with pdfplumber.open(file) as pdf:
            total_pages = len(pdf.pages)

            for page_num, page in enumerate(pdf.pages, 1):
                # ── Extract raw text ──
                raw_text = page.extract_text() or ""

                # ── Extract tables and convert to readable text ──
                table_text = ""
                try:
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            cleaned = [str(cell).strip() if cell else "" for cell in row]
                            if any(cleaned):
                                table_text += "  |  ".join(cleaned) + "\n"
                except Exception:
                    pass

                combined = (raw_text + "\n" + table_text).strip()

                # ── No native text layer — try local OCR, else keep as an image ──
                if len(combined) < config.MIN_PAGE_TEXT_CHARS:
                    if light:
                        scanned_pages += 1
                        continue

                    page_image = page.to_image(
                        resolution=config.OCR_RENDER_RESOLUTION_DPI
                    ).original

                    # Cooperative yield — see config.CPU_YIELD_SECONDS. Placed
                    # right after the render (the first and one of the
                    # heaviest steps for this page) so pages needing the
                    # fallback path don't run their CPU-bound work back-to-back
                    # with no gap for other threads on constrained hosts.
                    time.sleep(config.CPU_YIELD_SECONDS)

                    if _ocr_available() and not _page_has_annotation(page_image):
                        ocr_text = _ocr_page_text(page_image)
                        if len(ocr_text) >= config.MIN_PAGE_TEXT_CHARS:
                            combined = ocr_text
                            ocr_pages += 1
                        else:
                            scanned_pages += 1
                            fallback_images.append((page_num, _encode_page_jpeg(page_image)))
                            continue
                    else:
                        scanned_pages += 1
                        fallback_images.append((page_num, _encode_page_jpeg(page_image)))
                        continue

                # ── Deduplicate exact pages ──
                page_hash = hashlib.md5(combined.encode("utf-8")).hexdigest()
                if page_hash in seen_hashes:
                    skipped_pages += 1
                    continue
                seen_hashes.add(page_hash)

                pages_text.append(combined)

        full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages_text)

        return {
            "success":         bool(full_text.strip()) or bool(fallback_images),
            "text":            full_text,
            "page_count":      total_pages,
            "skipped_pages":   skipped_pages,
            "scanned_pages":   scanned_pages,
            "ocr_pages":       ocr_pages,
            "fallback_images": fallback_images,
            "use_fallback":    bool(fallback_images),
        }

    except Exception as e:
        return {
            "success":         False,
            "text":            "",
            "page_count":      0,
            "skipped_pages":   0,
            "scanned_pages":   0,
            "ocr_pages":       0,
            "fallback_images": [],
            "use_fallback":    True,
            "error":           str(e),
        }


# ── Content-block building (shared by realtime + batch submission) ───────────

def build_file_content(f, log_fn=None) -> dict:
    """
    Builds Claude content blocks for a single uploaded file: one text block
    covering native-extracted + OCR'd pages, plus a text/image block pair per
    page that needs the image fallback (handwriting/stamps detected, local
    OCR unavailable, or OCR produced too little text).

    log_fn: optional callable(str) for per-line logging (batch mode).

    Returns:
        {
            "content":        list  — content blocks for this file (extraction prompt not included)
            "page_count":     int
            "fallback_pages": int   — pages sent as images
            "ocr_pages":      int   — pages recovered via local OCR
            "notes":          list[str]  — human-readable notes for the email/UI
        }
    """
    def log(msg):
        if log_fn:
            log_fn(msg)

    extraction = extract_text_from_pdf(f)
    page_count = extraction.get("page_count", 1) or 1
    content    = []
    notes      = []

    if extraction["text"].strip():
        header_notes = []
        if extraction["skipped_pages"] > 0:
            header_notes.append(f"{extraction['skipped_pages']} duplicate page(s) skipped")
        if extraction["ocr_pages"] > 0:
            header_notes.append(f"{extraction['ocr_pages']} page(s) read via local OCR")
        if extraction["scanned_pages"] > 0:
            header_notes.append(f"{extraction['scanned_pages']} page(s) sent as images below")

        header = f"=== FILE: {f.name} ==="
        if header_notes:
            header += f" [{', '.join(header_notes)}]"

        content.append({"type": "text", "text": header + "\n\n" + extraction["text"]})
        if header_notes:
            notes.append(f"{f.name}: {', '.join(header_notes)}")

        log(
            f"{f.name} → text extraction ({page_count} pages"
            + (f", {extraction['ocr_pages']} via local OCR" if extraction["ocr_pages"] else "")
            + (f", {extraction['skipped_pages']} dup pages skipped" if extraction["skipped_pages"] else "")
            + ")"
        )

    payload_bytes = sum(len(b["text"]) for b in content if b["type"] == "text")

    for page_num, jpeg_bytes in extraction.get("fallback_images", []):
        b64_data = base64.standard_b64encode(jpeg_bytes).decode("utf-8")
        page_header = f"=== FILE: {f.name} — PAGE {page_num} (image below) ==="
        content.append({"type": "text", "text": page_header})
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64_data},
        })
        payload_bytes += len(page_header) + len(b64_data)
        log(f"{f.name} page {page_num} → image fallback (handwriting/stamp detected, or OCR unavailable)")

    if not content:
        # Extraction produced nothing at all (e.g. pdfplumber couldn't open the
        # file) — last-resort fallback: send the whole file as-is.
        f.seek(0)
        pdf_bytes = f.read()
        b64_data  = base64.standard_b64encode(pdf_bytes).decode("utf-8")
        content.append({
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64_data},
            "title": f.name,
        })
        payload_bytes += len(b64_data)
        log(f"{f.name} → whole-file PDF fallback (extraction produced no usable content)")

    payload_mb = payload_bytes / 1024 / 1024
    if payload_mb > config.MAX_REQUEST_PAYLOAD_MB:
        raise ValueError(
            f"{f.name} is too large to process in one request "
            f"({payload_mb:.1f}MB of image/text content, limit is "
            f"{config.MAX_REQUEST_PAYLOAD_MB:.0f}MB) — this happens with scanned "
            f"files that have many pages needing image fallback (handwriting/stamps, "
            f"or local OCR unavailable). Try splitting this file into smaller uploads."
        )

    return {
        "content":        content,
        "page_count":     page_count,
        "fallback_pages": extraction.get("scanned_pages", 0),
        "ocr_pages":      extraction.get("ocr_pages", 0),
        "notes":          notes,
    }


def build_file_content_chunks(f, log_fn=None) -> dict:
    """
    Like build_file_content(), but splits a file's fallback-image pages
    across multiple content chunks (each becomes its own Batch API request)
    when there are more than config.MAX_FALLBACK_PAGES_PER_REQUEST of them.

    Native/OCR text (if any) goes in chunk 0 only — every fallback image
    still appears exactly once across all chunks, so there's no duplicate
    extraction risk from splitting; the existing invoice-number+line-detail
    dedup pass (deduplicate_items) is still there as a safety net regardless.

    Submitting each chunk as its own separate Batch API job (done by the
    caller, not here) gives genuine breathing room between chunks — each
    submission's network call is a real I/O wait, unlike an in-process
    sleep, which per-page cooperative yields alone weren't enough to
    guarantee on very CPU-constrained hosts.

    log_fn: optional callable(str) for per-line logging (batch mode).

    Returns:
        {
            "chunks":         list[list[content_block]]  — one or more chunks, each a full content list
            "page_count":     int
            "fallback_pages": int   — pages sent as images (across all chunks)
            "ocr_pages":      int   — pages recovered via local OCR
            "notes":          list[str]  — human-readable notes for the email/UI
        }
    """
    def log(msg):
        if log_fn:
            log_fn(msg)

    extraction = extract_text_from_pdf(f)
    page_count = extraction.get("page_count", 1) or 1
    notes      = []

    header_notes = []
    if extraction["skipped_pages"] > 0:
        header_notes.append(f"{extraction['skipped_pages']} duplicate page(s) skipped")
    if extraction["ocr_pages"] > 0:
        header_notes.append(f"{extraction['ocr_pages']} page(s) read via local OCR")
    if extraction["scanned_pages"] > 0:
        header_notes.append(f"{extraction['scanned_pages']} page(s) sent as images below")
    if header_notes:
        notes.append(f"{f.name}: {', '.join(header_notes)}")

    fallback_images = extraction.get("fallback_images", [])
    chunk_size      = max(1, config.MAX_FALLBACK_PAGES_PER_REQUEST)
    image_groups    = [
        fallback_images[i:i + chunk_size]
        for i in range(0, len(fallback_images), chunk_size)
    ] or [[]]  # always at least one group, even with zero fallback images

    chunks = []

    for chunk_idx, image_group in enumerate(image_groups):
        block         = []
        payload_bytes = 0

        if chunk_idx == 0 and extraction["text"].strip():
            header = f"=== FILE: {f.name} ==="
            if header_notes:
                header += f" [{', '.join(header_notes)}]"
            block.append({"type": "text", "text": header + "\n\n" + extraction["text"]})
            payload_bytes += len(header) + len(extraction["text"])
            log(
                f"{f.name} → text extraction ({page_count} pages"
                + (f", {extraction['ocr_pages']} via local OCR" if extraction["ocr_pages"] else "")
                + (f", {extraction['skipped_pages']} dup pages skipped" if extraction["skipped_pages"] else "")
                + ")"
            )

        for page_num, jpeg_bytes in image_group:
            b64_data = base64.standard_b64encode(jpeg_bytes).decode("utf-8")
            page_header = f"=== FILE: {f.name} — PAGE {page_num} (image below) ==="
            block.append({"type": "text", "text": page_header})
            block.append({
                "type": "image",
                "source": {"type": "base64", "media_type": "image/jpeg", "data": b64_data},
            })
            payload_bytes += len(page_header) + len(b64_data)
            log(
                f"{f.name} page {page_num} → image fallback "
                f"(chunk {chunk_idx + 1}/{len(image_groups)}, handwriting/stamp detected, "
                f"or OCR unavailable)"
            )

        if not block:
            continue

        payload_mb = payload_bytes / 1024 / 1024
        if payload_mb > config.MAX_REQUEST_PAYLOAD_MB:
            raise ValueError(
                f"{f.name} (chunk {chunk_idx + 1}/{len(image_groups)}) is too large to "
                f"process in one request ({payload_mb:.1f}MB, limit is "
                f"{config.MAX_REQUEST_PAYLOAD_MB:.0f}MB) — try lowering "
                f"MAX_FALLBACK_PAGES_PER_REQUEST."
            )

        chunks.append(block)

    if not chunks:
        # Extraction produced nothing at all (e.g. pdfplumber couldn't open the
        # file) — last-resort fallback: send the whole file as-is, as one chunk.
        f.seek(0)
        pdf_bytes = f.read()
        b64_data  = base64.standard_b64encode(pdf_bytes).decode("utf-8")
        chunks.append([{
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64_data},
            "title": f.name,
        }])
        log(f"{f.name} → whole-file PDF fallback (extraction produced no usable content)")

    return {
        "chunks":         chunks,
        "page_count":     page_count,
        "fallback_pages": extraction.get("scanned_pages", 0),
        "ocr_pages":      extraction.get("ocr_pages", 0),
        "skipped_pages":  extraction.get("skipped_pages", 0),
        "notes":          notes,
    }


def build_captured_pages_content(images: list, name: str = "Scanned pages", log_fn=None) -> dict:
    """
    Like build_file_content_chunks(), but for a list of raw captured page
    images (bytes, e.g. from st.camera_input()) instead of a PDF file. A
    camera capture has no text layer to look for, so this skips
    pdfplumber/text-extraction entirely and goes straight to the same
    per-page handwriting-check -> OCR-or-image-fallback -> JPEG-encode
    pipeline already built for scanned PDF pages, then groups fallback pages
    into the same MAX_FALLBACK_PAGES_PER_REQUEST-sized chunks.

    images: list of raw image bytes, one per captured page, in capture order.
    name: label used in content headers / logs (e.g. "Scan batch 14:32").
    log_fn: optional callable(str) for per-line logging (batch mode).

    Returns the same shape as build_file_content_chunks():
        {
            "chunks":         list[list[content_block]]
            "page_count":     int
            "fallback_pages": int   — pages sent as images
            "ocr_pages":      int   — pages recovered via local OCR
            "notes":          list[str]
        }
    """
    def log(msg):
        if log_fn:
            log_fn(msg)

    seen_hashes     = set()
    pages_text      = []
    fallback_images = []
    scanned_pages   = 0
    ocr_pages       = 0
    skipped_pages   = 0
    page_count      = len(images)

    for page_num, img_bytes in enumerate(images, 1):
        try:
            pil_image = Image.open(io.BytesIO(img_bytes))
            pil_image.load()
        except Exception:
            # Not a decodable image (corrupt capture) — nothing usable to
            # send either as text or as an image, so skip it rather than
            # forwarding unreadable bytes mislabeled as a JPEG.
            skipped_pages += 1
            log(f"{name} page {page_num} → skipped (unreadable capture)")
            continue

        if _ocr_available() and not _page_has_annotation(pil_image):
            ocr_text = _ocr_page_text(pil_image)
            time.sleep(config.CPU_YIELD_SECONDS)
            if len(ocr_text) >= config.MIN_PAGE_TEXT_CHARS:
                page_hash = hashlib.md5(ocr_text.encode("utf-8")).hexdigest()
                if page_hash not in seen_hashes:
                    seen_hashes.add(page_hash)
                    pages_text.append(ocr_text)
                    ocr_pages += 1
                else:
                    skipped_pages += 1
                continue

        scanned_pages += 1
        fallback_images.append((page_num, _encode_page_jpeg(pil_image)))
        time.sleep(config.CPU_YIELD_SECONDS)

    full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages_text)

    header_notes = []
    if skipped_pages > 0:
        header_notes.append(f"{skipped_pages} duplicate/unreadable page(s) skipped")
    if ocr_pages > 0:
        header_notes.append(f"{ocr_pages} page(s) read via local OCR")
    if scanned_pages > 0:
        header_notes.append(f"{scanned_pages} page(s) sent as images below")

    notes = []
    if header_notes:
        notes.append(f"{name}: {', '.join(header_notes)}")

    chunk_size   = max(1, config.MAX_FALLBACK_PAGES_PER_REQUEST)
    image_groups = [
        fallback_images[i:i + chunk_size]
        for i in range(0, len(fallback_images), chunk_size)
    ] or [[]]  # always at least one group, even with zero fallback images

    chunks = []

    for chunk_idx, image_group in enumerate(image_groups):
        block         = []
        payload_bytes = 0

        if chunk_idx == 0 and full_text.strip():
            header = f"=== {name} ==="
            if header_notes:
                header += f" [{', '.join(header_notes)}]"
            block.append({"type": "text", "text": header + "\n\n" + full_text})
            payload_bytes += len(header) + len(full_text)
            log(
                f"{name} → OCR text ({page_count} pages"
                + (f", {ocr_pages} via local OCR" if ocr_pages else "")
                + (f", {skipped_pages} pages skipped" if skipped_pages else "")
                + ")"
            )

        for page_num, jpeg_bytes in image_group:
            b64_data = base64.standard_b64encode(jpeg_bytes).decode("utf-8")
            page_header = f"=== {name} — PAGE {page_num} (image below) ==="
            block.append({"type": "text", "text": page_header})
            block.append({
                "type": "image",
                "source": {"type": "base64", "media_type": "image/jpeg", "data": b64_data},
            })
            payload_bytes += len(page_header) + len(b64_data)
            log(
                f"{name} page {page_num} → image fallback "
                f"(chunk {chunk_idx + 1}/{len(image_groups)}, handwriting/stamp detected, "
                f"or OCR unavailable)"
            )

        if not block:
            continue

        payload_mb = payload_bytes / 1024 / 1024
        if payload_mb > config.MAX_REQUEST_PAYLOAD_MB:
            raise ValueError(
                f"{name} (chunk {chunk_idx + 1}/{len(image_groups)}) is too large to "
                f"process in one request ({payload_mb:.1f}MB, limit is "
                f"{config.MAX_REQUEST_PAYLOAD_MB:.0f}MB) — try lowering "
                f"MAX_FALLBACK_PAGES_PER_REQUEST."
            )

        chunks.append(block)

    if not chunks:
        log(f"{name} → no usable pages captured")

    return {
        "chunks":         chunks,
        "page_count":     page_count,
        "fallback_pages": scanned_pages,
        "ocr_pages":      ocr_pages,
        "skipped_pages":  skipped_pages,
        "notes":          notes,
    }


# ── Duplicate invoice number detection ───────────────────────────────────────

def deduplicate_items(items: list) -> tuple:
    """
    Removes items with duplicate invoice numbers, keeping first occurrence.
    Only active when config.SKIP_DUPLICATE_INVOICE_NUMBERS is True.

    Returns:
        (deduplicated_items, list_of_warning_strings)
    """
    if not config.SKIP_DUPLICATE_INVOICE_NUMBERS:
        return items, []

    seen_line_keys = {}
    deduplicated   = []
    warnings       = []

    for item in items:
        inv_no = _normalize_invoice_no(item.get("invoice_no") or item.get("in") or "")
        gstin = str(item.get("gstin") or item.get("g") or "").strip().upper()
        vendor = item.get("party_name") or item.get("pn") or "Unknown vendor"

        # Skip blank or null invoice numbers from dedup check
        if not inv_no or inv_no.lower() in ("null", "none", ""):
            deduplicated.append(item)
            continue

        line_parts = [
            str(item.get("description") or item.get("d") or "").strip().upper(),
            str(item.get("hsn_code") or item.get("h") or "").strip().upper(),
            str(item.get("qty") or item.get("q") or "").strip().upper(),
            str(item.get("rate") or item.get("r") or "").strip().upper(),
            str(item.get("taxable_value") or item.get("tv") or "").strip().upper(),
            str(item.get("total_value") or item.get("t") or "").strip().upper(),
        ]

        if gstin and GSTIN_RE.fullmatch(gstin):
            invoice_identity = f"gstin:{gstin}|invoice:{inv_no}"
            label = f"GSTIN '{gstin}' + invoice number '{inv_no}'"
        else:
            invoice_identity = f"invoice:{inv_no}"
            label = f"invoice number '{inv_no}'"

        dedup_key = invoice_identity + "|line:" + "|".join(line_parts)

        if dedup_key not in seen_line_keys:
            seen_line_keys[dedup_key] = vendor
            deduplicated.append(item)
        else:
            warn = (
                f"Duplicate line item for {label} "
                f"(vendor: {vendor}) — skipped. "
                f"First occurrence kept from: {seen_line_keys[dedup_key]}"
            )
            warnings.append(warn)

    return deduplicated, warnings


# ── Cost calculation ──────────────────────────────────────────────────────────

def calculate_cost(input_tokens: int, output_tokens: int) -> dict:
    input_cost  = (input_tokens  / 1_000_000) * config.PRICE_INPUT_PER_MTOK
    output_cost = (output_tokens / 1_000_000) * config.PRICE_OUTPUT_PER_MTOK
    total_cost  = input_cost + output_cost
    return {
        "input_tokens":    input_tokens,
        "output_tokens":   output_tokens,
        "input_cost_usd":  round(input_cost,  6),
        "output_cost_usd": round(output_cost, 6),
        "total_cost_usd":  round(total_cost,  6),
    }


def format_cost_summary(cost: dict, mode: str, realtime_cost: dict = None) -> str:
    lines = [
        f"Processing Mode   : {mode}",
        f"Input tokens      : {cost['input_tokens']:,}",
        f"Output tokens     : {cost['output_tokens']:,}",
        f"Input cost        : ${cost['input_cost_usd']:.4f}",
        f"Output cost       : ${cost['output_cost_usd']:.4f}",
        f"Total cost        : ${cost['total_cost_usd']:.4f}",
    ]
    if mode == "Batch API" and realtime_cost:
        saving = realtime_cost["total_cost_usd"] - cost["total_cost_usd"]
        lines.append(f"Saved vs Real-time: ${saving:.4f}  (50% Batch discount)")
    return "\n".join(lines)


# ── JSON parsing ──────────────────────────────────────────────────────────────

def parse_json_response(raw_text: str, token_limit: int = None) -> list:
    """
    Parses the abbreviated JSON array returned by Claude.
    Expands abbreviated keys to full names for use in Excel.
    Strips markdown fences if present.

    token_limit: the max_tokens value actually used for this request, for the
    truncation error message — defaults to config.MAX_TOKENS (real-time path);
    the batch path passes config.BATCH_MAX_TOKENS since it uses a higher cap.
    """
    # Abbreviated key → full key mapping
    KEY_MAP = {
        "s":  "sr_no",
        "pn": "party_name",
        "g":  "gstin",
        "in": "invoice_no",
        "id": "invoice_date",
        "d":  "description",
        "q":  "qty",
        "r":  "rate",
        "tv": "taxable_value",
        "cg": "cgst",
        "sg": "sgst",
        "ig": "igst",
        "h":  "hsn_code",
        "t":  "total_value",
    }

    text = raw_text.strip()
    if text.startswith("```"):
        lines = text.splitlines()
        text  = "\n".join(
            line for line in lines
            if not line.strip().startswith("```")
        ).strip()

    try:
        data = json.loads(text)
        if not isinstance(data, list):
            raise ValueError("Expected a JSON array at top level.")

        # Expand abbreviated keys
        expanded = []
        for item in data:
            expanded.append({
                KEY_MAP.get(k, k): (v if v is not None else "")
                for k, v in item.items()
            })
        return expanded

    except json.JSONDecodeError as e:
        # Detect truncation — happens when Claude hits max_tokens mid-response
        truncated = (
            not text.rstrip().endswith("]")
            or text.count("{") != text.count("}")
        )
        if truncated:
            limit = token_limit if token_limit is not None else config.MAX_TOKENS
            raise ValueError(
                f"Output truncated — Claude hit the max_tokens limit ({limit} tokens). "
                f"The JSON was cut off mid-response. Try splitting your files into smaller batches."
            )
        raise ValueError(f"JSON parse error: {e}\n\nRaw text:\n{text[:500]}")


# ── Excel creation ────────────────────────────────────────────────────────────

HEADERS = [
    "Sr No", "Party Name", "GSTIN", "Invoice No", "Invoice Date",
    "Description of Item", "Qty", "Rate", "Taxable Value",
    "CGST", "SGST", "IGST", "HSN Code", "Total Value",
]

FIELD_KEYS = [
    "sr_no", "party_name", "gstin", "invoice_no", "invoice_date",
    "description", "qty", "rate", "taxable_value",
    "cgst", "sgst", "igst", "hsn_code", "total_value",
]

COL_WIDTHS = [6, 30, 22, 22, 13, 46, 6, 14, 15, 12, 12, 12, 11, 14]

# Columns written as real numbers (not text) so Excel can sum/average them.
AMOUNT_FIELD_KEYS = {"rate", "taxable_value", "cgst", "sgst", "igst", "total_value"}


def create_excel(items: list, dup_warnings: list = None) -> bytes:
    """
    Creates a formatted Excel file from extracted invoice items.
    Optionally adds a Warnings sheet if duplicate invoices were skipped.
    Returns file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Register"

    thin         = Side(style="thin", color="BBBBBB")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    alt_fill     = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    white_fill   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    warn_fill    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Header row ──
    for col, header in enumerate(HEADERS, 1):
        cell            = ws.cell(row=1, column=col, value=header)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center_align
        cell.border     = border
    ws.row_dimensions[1].height = 30

    # ── Data rows ──
    for r_idx, item in enumerate(items, 2):
        fill = alt_fill if r_idx % 2 == 0 else white_fill
        for c_idx, key in enumerate(FIELD_KEYS, 1):
            val  = item.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx)

            if key == "sr_no":
                num = _parse_amount_or_none(val)
                cell.value = int(num) if num is not None else (str(val) if val else "")
            elif key in AMOUNT_FIELD_KEYS:
                # Write as a real number (not text) so Excel's status bar can
                # sum/average the column — the currency symbol Claude returned
                # is kept via number_format instead of being embedded in the text.
                num = _parse_amount_or_none(val)
                if num is not None:
                    cell.value = num
                    symbol = _detect_currency_symbol(val) or "₹"
                    cell.number_format = f'"{symbol}"#,##0.00'
                else:
                    cell.value = str(val) if val else ""
            else:
                cell.value = str(val) if val else ""

            cell.font      = Font(name="Arial", size=9)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = left_align if c_idx in (2, 6) else center_align

    # ── Column widths ──
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Warnings sheet (if duplicates were skipped) ──
    if dup_warnings:
        ws2 = wb.create_sheet("Duplicate Warnings")
        ws2.column_dimensions["A"].width = 100
        ws2.cell(row=1, column=1, value="Duplicate Invoice Warnings").font = Font(
            name="Arial", bold=True, size=11, color="CC0000"
        )
        for i, warn in enumerate(dup_warnings, 2):
            cell      = ws2.cell(row=i, column=1, value=warn)
            cell.font = Font(name="Arial", size=9)
            cell.fill = warn_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



# ── Amount parsing (shared by Excel numeric cells + Tally XML) ────────────────

CURRENCY_SYMBOLS = ["₹", "Rs.", "Rs", "$", "€", "£"]


def _detect_currency_symbol(val) -> str:
    """Returns the first currency symbol found in val, or "" if none."""
    s = str(val or "")
    for ch in CURRENCY_SYMBOLS:
        if ch in s:
            return ch
    return ""


def _parse_amount_or_none(val):
    """
    Same stripping as _parse_amount, but returns None (not 0.0) when the value
    is missing or genuinely non-numeric — lets callers tell "no data" apart
    from a real zero, instead of collapsing both to 0.0.
    """
    if val is None or str(val).strip() == "":
        return None
    s = str(val)
    for ch in CURRENCY_SYMBOLS + [",", " "]:
        s = s.replace(ch, "")
    s = s.strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_amount(val) -> float:
    """
    Safely extracts a float from a value that may be a string like "Rs.1,234.56"
    or "₹1,234.56" or just "1234.56". Returns 0.0 if unparseable.
    """
    num = _parse_amount_or_none(val)
    return num if num is not None else 0.0


# ── Tally XML generation ──────────────────────────────────────────────────────


def _tally_date(date_str: str) -> str:
    """
    Converts various date formats to Tally's required YYYYMMDD format.
    Tries common Indian invoice date formats.
    Returns today's date as fallback.
    """
    from datetime import date as date_type
    import re

    if not date_str:
        return datetime.now().strftime("%Y%m%d")

    s = str(date_str).strip()

    formats = [
        "%d-%b-%Y",     # 26-Sep-2024
        "%d/%m/%Y",     # 26/09/2024
        "%d-%m-%Y",     # 26-09-2024
        "%Y-%m-%d",     # 2024-09-26
        "%b %d, %Y",    # Sep 26, 2024
        "%d %b %Y",     # 26 Sep 2024
        "%d-%b-%y",     # 26-Sep-24
        "%m/%d/%Y",     # 09/26/2024
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y%m%d")
        except ValueError:
            continue

    # Last resort — extract 4-digit year and return Jan 1 of that year
    m = re.search(r"(\d{4})", s)
    if m:
        return m.group(1) + "0101"
    return datetime.now().strftime("%Y%m%d")


def _escape_xml(val) -> str:
    """Escapes special XML characters in a string value."""
    if val is None:
        return ""
    s = str(val)
    s = s.replace("&",  "&amp;")
    s = s.replace("<",  "&lt;")
    s = s.replace(">",  "&gt;")
    s = s.replace('"',  "&quot;")
    s = s.replace("'", "&apos;")
    return s


def canonicalize_party_names(items: list) -> list:
    """
    Returns a new list of items with party_name normalized so the same
    real-world vendor always gets exactly one ledger name across a batch —
    Claude's extraction of a vendor name isn't guaranteed byte-identical
    across different invoices from the same vendor (e.g. "Sri Venkateswara
    Filling Station" vs "SRI VENKATESWARA FILLING STATION"), and Tally
    matches ledger names by exact string. Without this, create_tally_
    ledger_masters_xml() and create_tally_xml() could disagree on the name
    for the same vendor, or the masters file could create two ledgers for
    one real vendor.

    The first-seen variant (whitespace-collapsed) becomes the canonical
    name for every item that matches on a case/whitespace-insensitive key.
    Must be applied identically to whatever gets passed to BOTH Tally file
    generators — this only touches the copy handed to them, not all_items
    used for the Excel register, which should still reflect exactly what
    was extracted.
    """
    canonical_by_key = {}
    normalized_items = []
    for item in items:
        raw = re.sub(r"\s+", " ", str(item.get("party_name") or "")).strip()
        key = raw.casefold()
        if key and key not in canonical_by_key:
            canonical_by_key[key] = raw
        new_item = dict(item)
        if key:
            new_item["party_name"] = canonical_by_key[key]
        normalized_items.append(new_item)
    return normalized_items


def tally_excluded_items(items: list) -> list:
    """
    Returns invoice_no/party_name/total_value for items excluded from both
    Tally files because total_value <= 0 (credit notes, corrections, or a
    zero-value extraction result) — create_tally_xml() and
    create_tally_ledger_masters_xml() both silently skip these rather than
    guessing at how to represent them (e.g. a credit note needs its debit/
    credit sides reversed, which risks compounding an extraction error into
    a wrong reversing entry), so this lets callers surface them as an
    explicit warning instead of losing them without a trace.
    """
    excluded = []
    for item in items:
        total = _parse_amount(item.get("total_value"))
        if total <= 0:
            excluded.append({
                "invoice_no":   item.get("invoice_no", ""),
                "party_name":   item.get("party_name", ""),
                "total_value":  total,
            })
    return excluded


def _build_voucher_xml(item: dict, tally_version: str) -> str:
    """
    Builds a single <VOUCHER> XML block for one invoice line item, as a
    Journal voucher — matching how these entries are actually recorded in
    practice (verified against real Tally examples), not an invoice-style
    Purchase voucher. Journal vouchers don't carry invoice-mode display
    fields (PARTYGSTIN, BASICBASEPARTYNAME, an "Invoice Voucher View"
    PERSISTEDVIEW) — the GSTIN is still preserved in the narration text.

    tally_version: "erp9" or "prime"

    Both use the same core schema — TallyPrime adds a RESERVEDNAME attribute
    which ERP 9 ignores, so we include it in both for safety.
    """
    date      = _tally_date(item.get("invoice_date", ""))
    party     = _escape_xml(item.get("party_name", ""))
    inv_no    = _escape_xml(item.get("invoice_no",  ""))
    # Unique per vendor+invoice+date, not invoice number alone — invoice
    # numbers aren't globally unique (two different vendors can both send an
    # "INV-001"), and REMOTEID/GUID are Tally's own keys for detecting "is
    # this the same object already imported" — a collision across vendors
    # would make Tally treat an unrelated invoice as an alteration of this one.
    voucher_key = hashlib.md5(
        f"{item.get('party_name','')}|{item.get('invoice_no','')}|{item.get('invoice_date','')}".encode("utf-8")
    ).hexdigest()
    # Bill-wise reference name — must be unique per party (Tally scopes bill
    # names to the ledger they're on, not globally), so falls back to a
    # voucher_key-derived value when the invoice number wasn't extracted,
    # rather than risking two bills on the same vendor both named "".
    bill_name = inv_no if item.get("invoice_no") else f"{date}-{voucher_key[:8]}"
    narration = _escape_xml(
        f"{item.get('description','')} | Invoice: {item.get('invoice_no','')} "
        f"| GSTIN: {item.get('gstin','') or 'N/A'} "
        f"| HSN: {item.get('hsn_code','') or 'N/A'}"
    )
    ledger    = _escape_xml(config.TALLY_DEFAULT_LEDGER)

    # Amounts
    total     = _parse_amount(item.get("total_value"))
    taxable   = _parse_amount(item.get("taxable_value")) or total
    cgst_amt  = _parse_amount(item.get("cgst"))
    sgst_amt  = _parse_amount(item.get("sgst"))
    igst_amt  = _parse_amount(item.get("igst"))

    # Determine GST type
    has_igst  = igst_amt > 0
    has_cgst  = cgst_amt > 0
    has_sgst  = sgst_amt > 0

    # Build ledger entries
    # Credit: Party ledger (creditor — we owe them)
    # Debit:  Expense ledger + GST ledgers
    entries = []

    # Debit — main expense ledger
    entries.append(f"""
        <ALLLEDGERENTRIES.LIST>
            <LEDGERNAME>{ledger}</LEDGERNAME>
            <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
            <AMOUNT>-{taxable:.2f}</AMOUNT>
            <GODOWNENTRIES.LIST/>
            <CATEGORYENTRIES.LIST/>
        </ALLLEDGERENTRIES.LIST>""")

    # Debit — GST ledgers
    if has_igst:
        entries.append(f"""
        <ALLLEDGERENTRIES.LIST>
            <LEDGERNAME>IGST</LEDGERNAME>
            <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
            <AMOUNT>-{igst_amt:.2f}</AMOUNT>
            <GODOWNENTRIES.LIST/>
            <CATEGORYENTRIES.LIST/>
        </ALLLEDGERENTRIES.LIST>""")
    if has_cgst:
        entries.append(f"""
        <ALLLEDGERENTRIES.LIST>
            <LEDGERNAME>CGST</LEDGERNAME>
            <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
            <AMOUNT>-{cgst_amt:.2f}</AMOUNT>
            <GODOWNENTRIES.LIST/>
            <CATEGORYENTRIES.LIST/>
        </ALLLEDGERENTRIES.LIST>""")
    if has_sgst:
        entries.append(f"""
        <ALLLEDGERENTRIES.LIST>
            <LEDGERNAME>SGST/UTGST</LEDGERNAME>
            <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
            <AMOUNT>-{sgst_amt:.2f}</AMOUNT>
            <GODOWNENTRIES.LIST/>
            <CATEGORYENTRIES.LIST/>
        </ALLLEDGERENTRIES.LIST>""")

    # Credit — party (sundry creditor), with bill-wise allocation so Tally
    # tracks this invoice as its own trackable bill under the vendor's
    # ledger, rather than folding it into one lump running balance — see
    # docs/tally-xml-import-design.md for why this matters for AP tracking.
    # BILLTYPE is always "New Ref": this app only records purchases, never
    # payments, so every bill it creates is a fresh one, not a settlement
    # against an existing bill ("Agst Ref").
    entries.append(f"""
        <ALLLEDGERENTRIES.LIST>
            <LEDGERNAME>{party}</LEDGERNAME>
            <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
            <AMOUNT>{total:.2f}</AMOUNT>
            <BILLALLOCATIONS.LIST>
                <NAME>{bill_name}</NAME>
                <BILLTYPE>New Ref</BILLTYPE>
                <AMOUNT>{total:.2f}</AMOUNT>
            </BILLALLOCATIONS.LIST>
            <GODOWNENTRIES.LIST/>
            <CATEGORYENTRIES.LIST/>
        </ALLLEDGERENTRIES.LIST>""")

    entries_xml = "".join(entries)

    # TallyPrime-specific attributes
    prime_attrs = ' RESERVEDNAME=""' if tally_version == "prime" else ""

    voucher = f"""
        <VOUCHER REMOTEID="{voucher_key}" VCHTYPE="Journal" ACTION="Create"{prime_attrs}>
        <DATE>{date}</DATE>
        <GUID>{voucher_key}</GUID>
        <VOUCHERTYPENAME>Journal</VOUCHERTYPENAME>
        <VOUCHERNUMBER>{inv_no}</VOUCHERNUMBER>
        <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>
        <ISINVOICE>No</ISINVOICE>
        <NARRATION>{narration}</NARRATION>{entries_xml}
        </VOUCHER>"""

    return voucher


def create_tally_xml(items: list, tally_version: str) -> bytes:
    """
    Generates a Tally-importable XML file from extracted invoice items.

    tally_version: "erp9"  → Tally ERP 9 format
                   "prime" → TallyPrime (3.x) format

    Both use the same core ENVELOPE/BODY/IMPORTDATA schema.
    TallyPrime adds minor attributes ERP 9 safely ignores.

    Returns XML as bytes.
    """
    version_comment = (
        "Tally ERP 9" if tally_version == "erp9"
        else "TallyPrime 3.x"
    )
    tally_messages = "".join(
        f"""
            <TALLYMESSAGE xmlns:UDF="TallyUDF">{_build_voucher_xml(item, tally_version)}
            </TALLYMESSAGE>"""
        for item in items
        if _parse_amount(item.get("total_value")) > 0
    )

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<!-- Tally Import File — {version_comment} -->
<!-- Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} -->
<!-- Default ledger: {_escape_xml(config.TALLY_DEFAULT_LEDGER)} -->
<!-- Reassign ledgers inside Tally after import as needed -->
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Import Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <IMPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>Vouchers</REPORTNAME>
            </REQUESTDESC>
            <REQUESTDATA>{tally_messages}
            </REQUESTDATA>
        </IMPORTDATA>
    </BODY>
</ENVELOPE>"""

    return xml.encode("utf-8")


def create_tally_ledger_masters_xml(items: list, tally_version: str) -> bytes:
    """
    Generates a Tally-importable XML file that creates every ledger master
    the vouchers from create_tally_xml() will reference: each unique
    party/vendor, the tax ledgers actually used in this batch (CGST/SGST/
    IGST), and the single default expense ledger. Tally's voucher import
    only references ledgers by name — it doesn't create them — so this file
    is meant to be imported FIRST, before the vouchers file, so every name
    the vouchers reference already exists.

    Uses the current single config.TALLY_DEFAULT_LEDGER for every vendor's
    expense line, matching create_tally_xml()'s current behavior. Real-world
    usage varies the expense category per vendor (e.g. "Repair and
    Maintenance" vs. "Diesel and Petrol") — not captured here yet; see
    docs/tally-xml-import-design.md.

    Vendors already existing in the accountant's Tally company are handled
    by Tally's own duplicate-ledger behavior on import, not detected here —
    every batch's unique parties are always included, so this file is safe
    to import repeatedly across batches.

    tally_version: "erp9" or "prime" — kept for interface symmetry with
    create_tally_xml(); ledger-master creation doesn't currently need any
    version-specific attributes.

    Returns XML as bytes.
    """
    version_comment = (
        "Tally ERP 9" if tally_version == "erp9"
        else "TallyPrime 3.x"
    )

    seen_parties  = set()
    party_ledgers = []
    has_igst = has_cgst = has_sgst = False

    for item in items:
        if _parse_amount(item.get("total_value")) <= 0:
            continue
        party = _escape_xml(item.get("party_name", ""))
        if party and party not in seen_parties:
            seen_parties.add(party)
            party_ledgers.append(party)
        if _parse_amount(item.get("igst")) > 0:
            has_igst = True
        if _parse_amount(item.get("cgst")) > 0:
            has_cgst = True
        if _parse_amount(item.get("sgst")) > 0:
            has_sgst = True

    def ledger_block(name: str, parent: str) -> str:
        return f"""
            <LEDGER Action="Create">
                <NAME>{name}</NAME>
                <PARENT>{parent}</PARENT>
                <OPENINGBALANCE>0</OPENINGBALANCE>
            </LEDGER>"""

    ledger_blocks = [ledger_block(_escape_xml(config.TALLY_DEFAULT_LEDGER), "Indirect Expenses")]
    if has_igst:
        ledger_blocks.append(ledger_block("IGST", "Duties &amp; Taxes"))
    if has_cgst:
        ledger_blocks.append(ledger_block("CGST", "Duties &amp; Taxes"))
    if has_sgst:
        ledger_blocks.append(ledger_block("SGST/UTGST", "Duties &amp; Taxes"))
    for party in party_ledgers:
        ledger_blocks.append(ledger_block(party, "Sundry Creditors"))

    tally_messages = "".join(
        f"""
            <TALLYMESSAGE xmlns:UDF="TallyUDF">{block}
            </TALLYMESSAGE>"""
        for block in ledger_blocks
    )

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<!-- Tally Ledger Masters Import File — {version_comment} -->
<!-- Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} -->
<!-- Import this file FIRST, before the Vouchers import file, so every -->
<!-- ledger name the vouchers reference already exists. -->
<ENVELOPE>
    <HEADER>
        <TALLYREQUEST>Import Data</TALLYREQUEST>
    </HEADER>
    <BODY>
        <IMPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>All Masters</REPORTNAME>
            </REQUESTDESC>
            <REQUESTDATA>{tally_messages}
            </REQUESTDATA>
        </IMPORTDATA>
    </BODY>
</ENVELOPE>"""

    return xml.encode("utf-8")



# ── Email ─────────────────────────────────────────────────────────────────────

def send_email(
    excel_bytes:       bytes,
    cost:              dict,
    mode:              str,
    file_count:        int,
    item_count:        int,
    user_email:        str   = None,
    dup_warnings:      list  = None,
    upload_dup_warnings: list = None,
    realtime_cost:     dict  = None,
    batch_id:          str   = None,
    tally_erp9_bytes:  bytes = None,
    tally_prime_bytes: bytes = None,
    tally_erp9_masters_bytes:  bytes = None,
    tally_prime_masters_bytes: bytes = None,
    tally_excluded:    list  = None,
) -> tuple:
    """
    Sends Excel + both Tally XML files as email attachments via Resend API.
    Recipients:
      - user_email  : the logged-in user who triggered the job (always included)
      - ADMIN_EMAIL : admin address(es) from config (always included)
    Uses HTTPS (port 443) — works on all hosting platforms including Render free tier.
    Returns (success: bool, message: str)
    """
    resend.api_key = config.RESEND_API_KEY

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    filename  = f"Invoice_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    subject   = f"Invoice Register Ready - {file_count} file(s) | {timestamp}"

    upload_dup_section = ""
    if upload_dup_warnings:
        upload_dup_section = (
            "\n-- Upload Duplicate PDF Warnings --\n"
            + "\n".join(
                f"  * {d.get('name')} skipped: same GSTIN + invoice number "
                f"as {d.get('duplicate_of')} "
                f"(GSTIN: {d.get('gstin')}, Invoice: {d.get('invoice_no')})"
                for d in upload_dup_warnings
            )
            + "\n"
        )

    dup_section = ""
    if dup_warnings:
        dup_section = (
            "\n-- Extracted Duplicate Invoice Warnings --\n"
            + "\n".join(f"  * {w}" for w in dup_warnings)
            + "\n"
        )

    tally_excluded_section = ""
    if tally_excluded:
        tally_excluded_section = (
            "\n-- Excluded From Tally Files (zero/negative total) --\n"
            + "\n".join(
                f"  * {x.get('party_name') or 'Unknown vendor'} / "
                f"Invoice {x.get('invoice_no') or 'N/A'}: "
                f"{x.get('total_value'):.2f} — likely a credit note/correction; "
                f"still in the Excel register, but not auto-posted to Tally — "
                f"enter manually if valid"
                for x in tally_excluded
            )
            + "\n"
        )

    body = (
        f"Hi,\n\n"
        f"Your invoice extraction is complete.\n\n"
        f"-- Summary --\n"
        f"Files processed                : {file_count}\n"
        f"Upload duplicate PDFs skipped  : {len(upload_dup_warnings or [])}\n"
        f"Extracted duplicates skipped   : {len(dup_warnings or [])}\n"
        f"Line items extracted           : {item_count}\n"
        f"Processed at                   : {timestamp}\n"
        + (f"Batch ID             : {batch_id}\n" if batch_id else "")
        + upload_dup_section
        + dup_section
        + tally_excluded_section
        + f"\n-- Note --\n"
        f"Attachments:\n"
        f"  Invoice_Register.xlsx        — full register for review\n"
        + (
            "\n  For Tally ERP 9, import in this order:\n"
            "    1. Tally_ERP9_LedgerMasters.xml  — creates any missing ledgers\n"
            "    2. Tally_ERP9_Import.xml         — the actual entries\n"
            if tally_erp9_bytes else ""
        )
        + (
            "\n  For TallyPrime, import in this order:\n"
            "    1. Tally_Prime_LedgerMasters.xml — creates any missing ledgers\n"
            "    2. Tally_Prime_Import.xml        — the actual entries\n"
            if tally_prime_bytes else ""
        )
        + f"\nAll values extracted directly from source documents.\n"
        f"Default expense ledger used: {config.TALLY_DEFAULT_LEDGER}\n"
        f"Reassign ledgers inside Tally after import as needed.\n"
        f"Missing fields are left blank.\n"
        + (f"See 'Duplicate Warnings' sheet in Excel for extracted duplicate invoices.\n" if dup_warnings else "")
        + "\nInvoice Processor MVP\n"
    )

    # Build recipient list:
    #   - logged-in user always receives their own results
    #   - admin email(s) always receive a copy
    admin_emails = [r.strip() for r in config.ADMIN_EMAIL.split(",") if r.strip()]
    recipients   = [user_email.lower().strip()]

    # Resend requires attachments as base64 strings
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_b64   = base64.b64encode(excel_bytes).decode("utf-8")
    attachments = [{"filename": filename, "content": excel_b64}]

    if tally_erp9_masters_bytes:
        attachments.append({
            "filename": f"Tally_ERP9_LedgerMasters_{ts}.xml",
            "content":  base64.b64encode(tally_erp9_masters_bytes).decode("utf-8"),
        })
    if tally_erp9_bytes:
        attachments.append({
            "filename": f"Tally_ERP9_Import_{ts}.xml",
            "content":  base64.b64encode(tally_erp9_bytes).decode("utf-8"),
        })
    if tally_prime_masters_bytes:
        attachments.append({
            "filename": f"Tally_Prime_LedgerMasters_{ts}.xml",
            "content":  base64.b64encode(tally_prime_masters_bytes).decode("utf-8"),
        })
    if tally_prime_bytes:
        attachments.append({
            "filename": f"Tally_Prime_Import_{ts}.xml",
            "content":  base64.b64encode(tally_prime_bytes).decode("utf-8"),
        })

    try:
        params = {
            "from":        config.RESEND_SENDER,
            "to":          recipients,
            "subject":     subject,
            "text":        body,
            "attachments": attachments,
            "bcc":         admin_emails
        }
        response = resend.Emails.send(params)
        # Resend returns {"id": "..."} on success
        if response and response.get("id"):
            return True, filename
        else:
            return False, f"Resend returned unexpected response: {response}"
    except Exception:
        return False, traceback.format_exc()
